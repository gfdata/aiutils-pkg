# -*- coding: utf-8 -*-
# @time: 2022/4/20 9:35
# @Author：lhf
# ----------------------
"""
rqalpha 运行结果的通用存储方式
"""
import os
import json
from copy import deepcopy
from functools import lru_cache

import colorama
import pandas as pd
from openpyxl import load_workbook, Workbook

from aiutils.json_obj import jsonable_encoder
from aiutils.openpyxl_tools import openpyxl_df_clear_write
import warnings

_ = f'FutureWarning：{__name__} 新的项目不要使用此模块，改用saveResult区分不同存储方式'
warnings.warn(_, FutureWarning)
colorama.init(autoreset=True)
print(colorama.Fore.YELLOW + _)


def save_config_json(context, output_path):
    """ 存储rqalpha的context.config为json格式；目录 output_path，文件名 config_rq.json
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)
    file = os.path.join(output_path, 'config_rq.json')
    if os.path.isfile(file):
        os.remove(file)

    # rqalpha的config
    with open(file, 'w') as fp:
        temp = deepcopy(context.config.convert_to_dict())
        del temp['base']['trading_calendar']
        temp = jsonable_encoder(temp)
        json.dump(temp, fp)


def read_rq_excel(output_path, name, use_cache=True):
    """ 读取rqalpha的运行结果
    * 不同name涉及的dt字段和要求不一样，所以此处不做处理；
    * 由外部调用方进行 astype 以及 set_index；
    """
    _f = 'sys_analyser.xlsx'  # 与save函数定义的文件名相对应
    file = os.path.join(output_path, _f)
    file = os.path.abspath(file)
    if os.path.exists(file) and os.path.isfile(file):
        return read_common_excel(file, name, use_cache)
    else:
        raise ValueError(f'缺少文件{_f} {output_path}')


def read_common_excel(file, sheet_name, use_cache=True):
    """通用的读取excel数据的函数
    * 简单读取某个sheet，结果整理由外部调用方进行 astype 以及 set_index
    """
    file_t = os.path.getmtime(file)
    if use_cache:
        return _read_excel_cache(file, file_t, sheet_name)
    else:
        return _read_excel(file, file_t, sheet_name)


@lru_cache()
def _read_excel_cache(file, file_t, sheet_name):
    return _read_excel(file, file_t, sheet_name)


def _read_excel(file, file_t, sheet_name):
    # xlrd 1.2.0以下才支持xlsx，2.0不支持；pandas 1.3.0 可用openpyxl引擎
    try:
        import xlrd
        df = pd.read_excel(file, sheet_name=sheet_name)
    except Exception as e:
        df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
    return df


def save_rq_excel(rq: dict, output_path):
    """
    存储rqalpha的运行结果；run_func返回的数据结构；目录 output_path，文件名 sys_analyser.xlsx
    * 参考 rqalpha.mod.rqalpha_mod_sys_analyser.report.generate_report
    * run_func返回的rq['sys_analyser']才是对映的上面mod的result_dict
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    # sys_analyser
    file = os.path.join(output_path, 'sys_analyser.xlsx')
    if os.path.exists(file) and os.path.isfile(file):
        wb = load_workbook(file)  # 这种方式，excel存在时不会删除里面的其他元素
    else:
        wb = Workbook()
    temp = rq.get('sys_analyser', None)
    if temp:
        summary = pd.Series(temp['summary']).to_frame('summary')
        openpyxl_df_clear_write(wb, 'summary', summary.reset_index())  # 此函数不会保存index，所以要reset一下
        # mod_sys_analyser中会用到的key
        for name in ["portfolio", "stock_account", "future_account",
                     "stock_positions", "future_positions", "trades"]:
            try:
                df = temp[name]
            except KeyError:
                continue
            # replace all date in dataframe as string
            if df.index.name == "date":  # 例如portfolio按日统计；trades带时间部分
                df = df.reset_index()
                df["date"] = df["date"].apply(lambda x: x.strftime("%Y-%m-%d"))
                df = df.set_index("date")
            # index name 和cols重复时
            if df.index.name in df.columns:
                del df[df.index.name]

            openpyxl_df_clear_write(wb, sheet_name=name, df=df.reset_index())

        # 最后保存
        wb.save(file)
        wb.close()
