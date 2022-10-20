# -*- coding: utf-8 -*-
# @time: 2022/7/23 13:36
# @Author：lhf
# ----------------------
import os
import json
from copy import deepcopy
from functools import lru_cache
import pandas as pd
from openpyxl import load_workbook, Workbook

from aiutils.json_obj import jsonable_encoder
from aiutils.openpyxl_tools import openpyxl_df_clear_write

from .result_abs import ResultAbstract


class ResultFile(ResultAbstract):
    """文件存储方式"""

    def __init__(self, id_name, id_script, id_vars,
                 save_space):
        self.id_name = id_name
        self.id_script = id_script
        self.id_vars = id_vars
        # 其他规则
        self._save_space = save_space
        if not os.path.exists(self.output_path):
            os.makedirs(self.output_path)

    @property
    def output_path(self):
        """ 此处 output_path 有意义；作为实际存储目录 """
        p = os.path.join(
            self._save_space, "+".join([self.id_name, self.id_script, self.id_vars])  # 建议与StrategyIdentify保持一致
        )
        return p

    def read_common_df(self, file_name, sheet_name, use_cache=True):
        file = os.path.join(self.output_path, f'{file_name}.xlsx')
        file = os.path.abspath(file)
        if os.path.exists(file) and os.path.isfile(file):
            return read_common_excel(file, sheet_name, use_cache)
        else:
            raise ValueError(f'缺少文件{file}')

    # ------------------------------------------------------------------------------------
    def is_exist(self) -> bool:
        """通过判断context_config.json是否存在；因此self.save_context_config通常最后执行写入 """
        file = os.path.join(self.output_path, 'context_config.json')
        if os.path.isfile(file):
            return True
        else:
            return False

    def save_common_df(self, file_name, sheet_name, df, **kwargs):
        """通用的存储DataFrame；注意使用的openpyxl_df_clear_write不会保存df.index"""
        from openpyxl import load_workbook, Workbook
        from aiutils.openpyxl_tools import openpyxl_df_clear_write
        file = os.path.join(self.output_path, f'{file_name}.xlsx')
        if os.path.exists(file) and os.path.isfile(file):
            wb = load_workbook(file)
        else:
            wb = Workbook()
        openpyxl_df_clear_write(wb, sheet_name, df)
        # 保存并关闭
        wb.save(file)
        wb.close()

    def save_common_jsonable(self, name, obj):
        """ 通用的存储可json序列化的数据 """
        file = os.path.join(self.output_path, f'{name}.json')
        if os.path.isfile(file):
            os.remove(file)
        with open(file, 'w') as fp:
            temp = jsonable_encoder(obj)
            json.dump(temp, fp)

    def save_context_config(self, context):
        file = os.path.join(self.output_path, 'context_config.json')
        if os.path.isfile(file):
            os.remove(file)
        # rqalpha的config
        with open(file, 'w') as fp:
            temp = deepcopy(context.config.convert_to_dict())
            del temp['base']['trading_calendar']
            temp = jsonable_encoder(temp)
            json.dump(temp, fp)

    def save_rq_run(self, rq: dict):
        """
        存储rqalpha的运行结果；run_func返回的数据结构；目录 output_path，文件名 sys_analyser.xlsx
        * 参考 rqalpha.mod.rqalpha_mod_sys_analyser.report.generate_report
        """
        # 存储sys_analyser
        file = os.path.join(self.output_path, 'sys_analyser.xlsx')
        if os.path.exists(file) and os.path.isfile(file):
            wb = load_workbook(file)  # 这种方式，excel存在时不会删除里面的其他元素
        else:
            wb = Workbook()
        temp = rq.get('sys_analyser', None)
        if temp:
            summary = pd.Series(temp['summary']).to_frame('summary')
            summary=summary.astype('str')  # rqalpha4.11新增一项 回撤期 是IndexRange
            openpyxl_df_clear_write(wb, 'summary', summary.reset_index())  # 此函数不会保存index，所以要reset一下
            # mod_sys_analyser中会用到的key
            for name in ["portfolio", "stock_account", "future_account",
                         "stock_positions", "future_positions", "trades"]:
                try:
                    df = temp[name]
                except KeyError:
                    continue
                # replace all date in dataframe as string
                if df.index.name == "date":  # 例如portfolio按日统计；trades带时间部分
                    df = df.reset_index()
                    df["date"] = df["date"].apply(lambda x: x.strftime("%Y-%m-%d"))
                    df = df.set_index("date")
                # index name 和cols重复时
                if df.index.name in df.columns:
                    del df[df.index.name]
                openpyxl_df_clear_write(wb, sheet_name=name, df=df.reset_index())

            # 最后保存
            wb.save(file)
            wb.close()


def read_common_excel(file, sheet_name, use_cache=True):
    """通用的读取excel数据的函数
    * 简单读取某个sheet，结果整理由外部调用方进行 astype 以及 set_index
    """
    file_t = os.path.getmtime(file)
    if use_cache:
        return _read_excel_cache(file, file_t, sheet_name)
    else:
        return _read_excel(file, file_t, sheet_name)


@lru_cache()
def _read_excel_cache(file, file_t, sheet_name):
    return _read_excel(file, file_t, sheet_name)


def _read_excel(file, file_t, sheet_name):
    # xlrd 1.2.0以下才支持xlsx，2.0不支持；pandas 1.3.0 可用openpyxl引擎
    try:
        import xlrd
        df = pd.read_excel(file, sheet_name=sheet_name)
    except Exception as e:
        df = pd.read_excel(file, sheet_name=sheet_name, engine="openpyxl")
    return df
