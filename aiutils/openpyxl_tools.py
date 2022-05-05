# -*- coding: utf-8 -*-
# @time: 2022/3/24 11:03
# @Author：lhf
# ----------------------
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def openpyxl_df_remove_write(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> Workbook:
    """
    写入特点：
    * 删除该sheet_name再创建写入；单元格的值、其它元素都会删掉；
    * 统一忽略df的index，从Cell(1,1)也就是A1单元格，写入header和数据；
    """
    # sheet存在则删除
    if sheet_name in wb.get_sheet_names():
        try:
            wb.remove(wb[sheet_name])  # ws.remove传入对象而非str
        except Exception as e:
            print(f'删除表格失败 [{wb.path}][{sheet_name}] {e}')

    # 写入数据
    ws = wb.create_sheet(sheet_name)
    r_num = 1
    for r in dataframe_to_rows(df, index=False, header=True):
        for i in range(len(r)):
            ws.cell(r_num, i + 1).value = r[i]
        r_num += 1

    # 样式
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

    return wb


def openpyxl_df_clear_write(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> Workbook:
    """
    写入特点：
    * 清除sheet_name里面所有单元格的值；wb是修改模式时，保留其它元素例如图表、色阶；
    * 统一忽略df的index，从Cell(1,1)也就是A1单元格，写入header和数据；
    """
    # sheet是否存在
    if not sheet_name in wb.get_sheet_names():
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    # clear旧数据
    try:
        for col in ws.columns:
            for c in col:
                c.value = None
    except Exception as e:
        print(f'清空表格数据出错 [{wb.path}][{sheet_name}] {e}')

    # write新数据
    if df is None or df.empty:
        return wb

    r_num = 1  # 从1开始计数
    # dataframe_to_rows函数处理index=True有问题，不能对齐index.name到header；直接全设为False
    for r in dataframe_to_rows(df, index=False, header=True):
        # ws.append(r)  # append会在之后添加，而非覆盖重写
        for i in range(len(r)):
            ws.cell(r_num, i + 1).value = r[i]
        r_num += 1

    return wb
